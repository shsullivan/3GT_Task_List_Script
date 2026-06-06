"""
Create PM Task List from csv files in a provided path
"""

from enum import Enum
import json
import logging
from pathlib import Path
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

#-----------------------------------------------------
# Constants & Configurations
#-----------------------------------------------------

# Columns to keep from raw data
KEEP_COLUMNS = ["Task ID", "Task Name", "Assigned To", "Cell"]

# Frequently used file paths
CONFIG_DIR = Path(__file__).parent / "config"
TASK_DICT_PATH = CONFIG_DIR / "task_owners.csv"
SECTION_CONFIG_PATH = CONFIG_DIR / "section_config.json"
WEEKLY_EXPORTS = Path(__file__).parent / "weekly_exports"

# Formatting and styles for writing Excel file
HEADER_FONT = Font(bold=True, size=14, name="Arial Narrow", color="FFFFFF")
NOTE_FONT = Font(bold=True, size=12, name="Arial Narrow", color="FF0000")
TASK_FONT = Font(size=11)
CHECK_FONT = Font(bold=True, size=11)

TITLE_FILL = PatternFill(start_color="FF8EA9DB",end_color="FF8EA9DB",fill_type="solid")
HEADER_FILL = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
CHECK_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
THIN_SIDE = Side(style="thin")

CENTER = Alignment(horizontal="center", vertical="bottom")
LEFT = Alignment(horizontal="left", vertical="bottom")
WRAP = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

COLUMN_WIDTHS = {"A": 12.85, "B": 97.28, "C": 21.00, "D": 20.85}

HEADER_ROW_HEIGHT = 22
NOTE_ROW_HEIGHT = 18
TASK_ROW_HEIGHT = 15

#-----------------------------------------------------
# Enums and Custom Types
#-----------------------------------------------------

# Enum used to verify loaded taskDict has no errors at runtime
class TaskOwner(Enum):
    NIGHTS = "Night Shift Line Tech"
    DAYS = "Day Shift Line Tech"
    IMM = "IMM PM Tech"
    PMTEAM = "PM Team Tech"
    SUPPORT = "PM Support"
    CLEAN = "Cleaning Crew"
    VSPL = "VSPL"
    CAL = "Calibrations"
    MCKEN = "McKendrees"
    MILLER = "Miller Elec."
    RELIABILITY = "Reliability Tech"
    SME = "SME"
    SHUTDOWN = "Shutdown Item"

# List used to define task order on final sheet
CUSTOM_ORDER = [
    "",
    TaskOwner.NIGHTS.value,
    TaskOwner.PMTEAM.value,
    TaskOwner.DAYS.value,
    TaskOwner.VSPL.value,
    TaskOwner.SME.value,
    TaskOwner.SUPPORT.value,
    TaskOwner.CAL.value,
    TaskOwner.MILLER.value,
    TaskOwner.MCKEN.value,
    TaskOwner.RELIABILITY.value,
    TaskOwner.IMM.value,
    TaskOwner.CLEAN.value,
    TaskOwner.SHUTDOWN.value
]

# Section-specific sort rules for the final workbook.
SECTION_SORT_COLUMNS = {
    TaskOwner.IMM.value: ["Cell", "Task ID"]
}

#-----------------------------------------------------
# Logger Setup
#-----------------------------------------------------

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

#-----------------------------------------------------
# Data Loading
#-----------------------------------------------------

def load_task_dict(filepath: Path) -> dict[str, str]:
    """Load and validate task-to-owner mapping from csv file"""
    valid_owners = {member.value for member in TaskOwner}
    task_dict = {}

    with open(filepath, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            task_id = row["task_id"].strip()
            owner = row["owner"].strip()
            task_dict[task_id] = owner

    # CSV entry validation
    invalid = {k: v for k, v in task_dict.items() if v not in valid_owners}
    if invalid:
        raise ValueError(f"Uknown task owners in {filepath}: {invalid}")
    logger.info("Loaded %d task mappings from %s", len(task_dict), filepath.name)
    return task_dict

def load_csvs(directory: Path) -> pd.DataFrame:
    """Read and concatenate all CSVs in *directory*."""
    csv_files = list(directory.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {directory}")

    logger.info("Found %d CSV files", len(csv_files))
    return pd.concat(
        (pd.read_csv(f, header=0) for f in csv_files),
        ignore_index=True
    )

def load_section_config(filepath: Path) -> list[dict]:
    """Load section configuration JSON."""
    with open(filepath, "r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, list):
        raise ValueError("section_config.json must be a list of section objects")
    return data

#-----------------------------------------------------
# Data Transformation
#-----------------------------------------------------

def assign_owners(df: pd.DataFrame, task_dict: dict[str, str]) -> pd.DataFrame:
    """Populate the Assigned To column from the task lookup file"""
    df = df.copy()
    df["Assigned To"] = df["Task ID"].map(task_dict)
    unmapped = df["Assigned To"].isna().sum()
    if unmapped:
        logger.warning("%d tasks had no owner mapping", unmapped)
    return df

def sort_by_owner(df: pd.DataFrame, order: list[str]) -> pd.DataFrame:
    """Sort rows by custom owner order; unmapped owners go to the end."""
    owner_dtype = pd.CategoricalDtype(categories=order, ordered=True)
    df = df.copy()
    df["Assigned To"] = df["Assigned To"].astype(owner_dtype)
    return df.sort_values("Assigned To").reset_index(drop=True)

#-----------------------------------------------------
# Data/Config Validation
#-----------------------------------------------------
def get_configured_owners(section_config: list[dict]) -> set[str]:
    """Return every owner included in section_config.json"""
    configured_owners = set()

    for section in section_config:
        owner_config = section["owner"]
        owners = owner_config if isinstance(owner_config, list) else [owner_config]
        configured_owners.update(owners)

    return configured_owners

def warn_for_unwritten_tasks(df: pd.DataFrame, section_config: list[dict]) -> None:
    """Warn for unwritten tasks"""
    configured_owners = get_configured_owners(section_config)
    task_owners = set(df["Assigned To"].dropna().unique())

    missing_owners = task_owners - configured_owners

    if missing_owners:
        missing_tasks = df[df["Assigned To"].isin(missing_owners)]

        logger.warning(
            "%d tasks will not be written because their owners are missing from the section_config.json: %s",
            len(missing_tasks),
            sorted(missing_owners))

def validate_section_config(section_config: list[dict]) -> None:
    """Validate section_config.json owner values"""
    valid_owners = {member.value for member in TaskOwner}
    configured_owners = get_configured_owners(section_config)

    invalid_owners = configured_owners - valid_owners
    if invalid_owners:
        raise ValueError(f"Unknown owners in section_config.json: {sorted(invalid_owners)}")

def warn_for_duplicate_task_ids(df: pd.DataFrame) -> None:
    """Warn when duplicate task IDs are present in the export data"""
    duplicate_task_ids = sorted(df.loc[df["Task ID"].duplicated(), "Task ID"].dropna().unique())

    if duplicate_task_ids:
        logger.warning("Duplicate task IDs found in export data: %s", duplicate_task_ids)

#-----------------------------------------------------
# Excel Writing / Formatting (openpyxl)
#-----------------------------------------------------

def get_tasks_for_section(df: pd.DataFrame, owner_config) -> pd.DataFrame:
    owners = owner_config if isinstance(owner_config, list) else [owner_config]
    section_tasks = df[df["Assigned To"].isin(owners)][KEEP_COLUMNS]

    if len(owners) == 1 and owners[0] in SECTION_SORT_COLUMNS:
        return section_tasks.sort_values(
            by=SECTION_SORT_COLUMNS[owners[0]],
            na_position="last"
        )

    return section_tasks

def apply_group_border(ws, row_num: int, start_col: int = 1, end_col: int = 4) -> None:
    """Apply one outside border around a horizontal group of cells."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).border = Border(
            left=THIN_SIDE if col == start_col else None,
            right=THIN_SIDE if col == end_col else None,
            top=THIN_SIDE,
            bottom=THIN_SIDE
        )

def apply_group_border_range(
    ws,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int = 4
) -> None:
    """Apply one outside border around a rectangular group of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = Border(
                left=THIN_SIDE if col == start_col else None,
                right=THIN_SIDE if col == end_col else None,
                top=THIN_SIDE if row == start_row else None,
                bottom=THIN_SIDE if row == end_row else None
            )

def write_title_row(ws, row_num: int) -> int:
    ws.cell(row=row_num, column=2, value="TAM - 30 - 48 Week PM")
    ws.cell(row=row_num, column=3, value=pd.Timestamp.today().strftime("%B %d, %Y"))

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = TITLE_FILL
        cell.alignment = CENTER

    apply_group_border(ws, row_num)

    ws.row_dimensions[row_num].height = HEADER_ROW_HEIGHT
    return row_num + 1

def write_section_header(ws, row_num: int, header: str) -> int:
    """Write green section header row"""
    ws.cell(row=row_num, column=2, value=header)

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    apply_group_border(ws, row_num)

    ws.row_dimensions[row_num].height = HEADER_ROW_HEIGHT
    return row_num + 1

def write_notes(ws, row_num: int, notes: list[str]) -> int:
    """Write yellow notes rows"""
    if not notes:
        return row_num

    start_row = row_num

    for note in notes:
        ws.cell(row=row_num, column=2, value=note)

        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = NOTE_FILL
            cell.font = NOTE_FONT
            cell.alignment = WRAP_CENTER

        ws.row_dimensions[row_num].height = NOTE_ROW_HEIGHT
        row_num += 1

    apply_group_border_range(ws, start_row, row_num - 1)

    return row_num

def write_task_rows(ws, row_num: int, tasks: pd.DataFrame) -> int:
    """Write generated task rows from the CSV files"""
    for _, task in tasks.iterrows():
        values = [task[column] for column in KEEP_COLUMNS]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = TASK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_manual_tasks(ws, row_num: int, manual_tasks: list[list[str]]) -> int:
    """Write manually configured task rows from section_config.json"""
    for task in manual_tasks:
        for col_num, value in enumerate(task, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = TASK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_check_rows(ws, row_num: int, checks: list[list[str]]) -> int:
    """Write red process-confirmation check rows"""
    for check in checks:
        for col_num, value in enumerate(check, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = CHECK_FILL
            cell.font = CHECK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_section(ws, row_num: int, section: dict, df: pd.DataFrame) -> int:
    section_tasks = get_tasks_for_section(df, section["owner"])

    row_num = write_section_header(ws, row_num, section["header"])
    row_num = write_notes(ws, row_num, section.get("notes", []))
    row_num = write_task_rows(ws, row_num, section_tasks)
    row_num = write_manual_tasks(ws, row_num, section.get("manual_tasks", []))
    row_num = write_check_rows(ws, row_num, section.get("checks", []))

    return row_num

def set_column_widths(ws) -> None:
    """Apply column widths to the Task List sheet."""
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

#-----------------------------------------------------
# I/O Helpers
#-----------------------------------------------------

def get_output_filename() -> str:
    """Prompt until a valid filename is entered."""
    while True:
        name = input("Enter new task list filename: ").strip()
        if name.endswith(".xlsx"):
            return name
        print("Filename must end with .xlsx")

def write_workbook(df: pd.DataFrame, section_config: list[dict], output_path: Path) -> None:
    """Write sectioned Task List workbook to *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task List"

    row_num = 1

    row_num = write_title_row(ws, row_num)

    for section in section_config:
        row_num = write_section(ws, row_num, section, df)

    set_column_widths(ws)

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)

def main() -> None:
    path = WEEKLY_EXPORTS
    if not path.exists():
        raise NotADirectoryError(f"Invalid directory path {path}")

    task_dict = load_task_dict(TASK_DICT_PATH)
    section_config = load_section_config(SECTION_CONFIG_PATH)
    validate_section_config(section_config)

    df = load_csvs(path)
    df = assign_owners(df, task_dict)
    df = sort_by_owner(df, CUSTOM_ORDER)

    warn_for_duplicate_task_ids(df)
    warn_for_unwritten_tasks(df, section_config)

    output_filename = get_output_filename()
    write_workbook(df, section_config,path / output_filename)


if __name__ == "__main__":
    main()
