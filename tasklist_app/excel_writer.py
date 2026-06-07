#-----------------------------------------------------
# Excel Writing / Formatting (openpyxl)
#-----------------------------------------------------

from datetime import datetime
import logging
from pathlib import Path
import openpyxl
import pandas as pd
from openpyxl.styles import Border
from tasklist_app.constants import (
    KEEP_COLUMNS, SECTION_SORT_COLUMNS, THIN_SIDE,
    HEADER_FONT, TITLE_FILL, CENTER, HEADER_ROW_HEIGHT,
    HEADER_FILL, NOTE_FILL, NOTE_FONT, WRAP_CENTER,
    NOTE_ROW_HEIGHT, TASK_FONT, WRAP, THIN_BORDER,
    TASK_ROW_HEIGHT, CHECK_FILL, CHECK_FONT, COLUMN_WIDTHS
)

logger = logging.getLogger(__name__)

def get_tasks_for_section(df: pd.DataFrame, owner_config) -> pd.DataFrame:
    owners = owner_config if isinstance(owner_config, list) else [owner_config]
    section_tasks = df[df["Assigned To"].isin(owners)][KEEP_COLUMNS]

    if len(owners) == 1 and owners[0] in SECTION_SORT_COLUMNS:
        return section_tasks.sort_values(
            by=SECTION_SORT_COLUMNS[owners[0]],
            na_position="last"
        )

    return section_tasks

def apply_group_border(ws, row_num: int, start_col: int = 1, end_col: int = 4) -> None:
    """Apply one outside border around a horizontal group of cells."""
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).border = Border(
            left=THIN_SIDE if col == start_col else None,
            right=THIN_SIDE if col == end_col else None,
            top=THIN_SIDE,
            bottom=THIN_SIDE
        )

def apply_group_border_range(
    ws,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int = 4
) -> None:
    """Apply one outside border around a rectangular group of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = Border(
                left=THIN_SIDE if col == start_col else None,
                right=THIN_SIDE if col == end_col else None,
                top=THIN_SIDE if row == start_row else None,
                bottom=THIN_SIDE if row == end_row else None
            )

def write_title_row(ws, row_num: int, title: str, display_date: str) -> int:
    ws.cell(row=row_num, column=2, value=title)
    ws.cell(row=row_num, column=3, value=display_date)

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = TITLE_FILL
        cell.alignment = CENTER

    apply_group_border(ws, row_num)

    ws.row_dimensions[row_num].height = HEADER_ROW_HEIGHT
    return row_num + 1

def write_section_header(ws, row_num: int, header: str) -> int:
    """Write green section header row"""
    ws.cell(row=row_num, column=2, value=header)

    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    apply_group_border(ws, row_num)

    ws.row_dimensions[row_num].height = HEADER_ROW_HEIGHT
    return row_num + 1

def write_notes(ws, row_num: int, notes: list[str]) -> int:
    """Write yellow notes rows"""
    if not notes:
        return row_num

    start_row = row_num

    for note in notes:
        ws.cell(row=row_num, column=2, value=note)

        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = NOTE_FILL
            cell.font = NOTE_FONT
            cell.alignment = WRAP_CENTER

        ws.row_dimensions[row_num].height = NOTE_ROW_HEIGHT
        row_num += 1

    apply_group_border_range(ws, start_row, row_num - 1)

    return row_num

def write_task_rows(ws, row_num: int, tasks: pd.DataFrame) -> int:
    """Write generated task rows from the CSV files"""
    for _, task in tasks.iterrows():
        values = [task[column] for column in KEEP_COLUMNS]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = TASK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_manual_tasks(ws, row_num: int, manual_tasks: list[list[str]]) -> int:
    """Write manually configured task rows from section_config.json"""
    for task in manual_tasks:
        for col_num, value in enumerate(task, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = TASK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_check_rows(ws, row_num: int, checks: list[list[str]]) -> int:
    """Write red process-confirmation check rows"""
    for check in checks:
        for col_num, value in enumerate(check, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = CHECK_FILL
            cell.font = CHECK_FONT
            cell.alignment = WRAP if col_num == 2 else CENTER
            cell.border = THIN_BORDER

        ws.row_dimensions[row_num].height = TASK_ROW_HEIGHT
        row_num += 1

    return row_num

def write_section(ws, row_num: int, section: dict, df: pd.DataFrame) -> int:
    section_tasks = get_tasks_for_section(df, section["owner"])

    row_num = write_section_header(ws, row_num, section["header"])
    row_num = write_notes(ws, row_num, section.get("notes", []))
    row_num = write_task_rows(ws, row_num, section_tasks)
    row_num = write_manual_tasks(ws, row_num, section.get("manual_tasks", []))
    row_num = write_check_rows(ws, row_num, section.get("checks", []))

    return row_num

def set_column_widths(ws) -> None:
    """Apply column widths to the Task List sheet."""
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def write_workbook(
        df: pd.DataFrame,
        section_config: list[dict],
        output_path: Path,
        title: str,
        display_date: str
    ) -> None:

    """Write sectioned Task List workbook to *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task List"

    row_num = 1

    row_num = write_title_row(ws, row_num, title, display_date)

    for section in section_config:
        row_num = write_section(ws, row_num, section, df)

    set_column_widths(ws)

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)