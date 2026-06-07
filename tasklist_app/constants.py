#-----------------------------------------------------
# Constants & Configurations
#-----------------------------------------------------

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
import sys
from tasklist_app.models import TaskOwner

def get_project_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent.parent

# Columns to keep from raw data
KEEP_COLUMNS = ["Task ID", "Task Name", "Assigned To", "Cell"]

# Frequently used file paths
PROJECT_ROOT = get_project_root()
CONFIG_DIR = PROJECT_ROOT / "config"
TASK_DICT_PATH = CONFIG_DIR / "task_owners.csv"
SECTION_CONFIG_PATH = CONFIG_DIR / "section_config.json"
WEEKLY_EXPORTS = PROJECT_ROOT / "weekly_exports"

# Formatting and styles for writing Excel file
HEADER_FONT = Font(bold=True, size=14, name="Arial Narrow", color="FFFFFF")
NOTE_FONT = Font(bold=True, size=12, name="Arial Narrow", color="FF0000")
TASK_FONT = Font(size=11)
CHECK_FONT = Font(bold=True, size=11)

TITLE_FILL = PatternFill(start_color="FF8EA9DB",end_color="FF8EA9DB",fill_type="solid")
HEADER_FILL = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
CHECK_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
THIN_SIDE = Side(style="thin")

CENTER = Alignment(horizontal="center", vertical="bottom")
LEFT = Alignment(horizontal="left", vertical="bottom")
WRAP = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

COLUMN_WIDTHS = {"A": 12.85, "B": 97.28, "C": 21.00, "D": 20.85}

HEADER_ROW_HEIGHT = 22
NOTE_ROW_HEIGHT = 18
TASK_ROW_HEIGHT = 15

# List used to define task order on final sheet
CUSTOM_ORDER = [
    "",
    TaskOwner.NIGHTS.value,
    TaskOwner.PMTEAM.value,
    TaskOwner.DAYS.value,
    TaskOwner.VSPL.value,
    TaskOwner.SME.value,
    TaskOwner.SUPPORT.value,
    TaskOwner.CAL.value,
    TaskOwner.MILLER.value,
    TaskOwner.MCKEN.value,
    TaskOwner.RELIABILITY.value,
    TaskOwner.IMM.value,
    TaskOwner.CLEAN.value,
    TaskOwner.SHUTDOWN.value
]

# Section-specific sort rules for the final workbook.
SECTION_SORT_COLUMNS = {
    TaskOwner.IMM.value: ["Cell", "Task ID"]
}
